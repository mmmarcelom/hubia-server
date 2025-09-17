"""
Processador de documentos para sumarização usando Ollama Gemma
"""
import httpx
from typing import Dict, Any, List
from .base64_processor import Base64Processor
# TaskType enum values
SUMMARIZE = "summarize"

# Document processing libraries
import PyPDF2
from docx import Document
import openpyxl
from io import BytesIO

class DocumentProcessor(Base64Processor):
    """Processador de documentos para sumarização"""
    
    def __init__(self, config):
        super().__init__(config)
        self.allowed_mime_types = [
            "application/pdf", "application/msword", 
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "text/plain", "text/csv"
        ]
        self.allowed_extensions = ["pdf", "doc", "docx", "xls", "xlsx", "txt", "csv"]
    
    def _prepare_document_data(self, content: str) -> Dict[str, Any]:
        """Prepara dados de documento a partir de data URL base64 ou texto"""
        try:
            # Check if it's a data URL
            if content.startswith('data:'):
                header, data = content.split(',', 1)
                mime_type = header.split(';')[0].replace('data:', '')
                file_data = data
                
                # Decode base64 to get file size
                import base64
                file_bytes = base64.b64decode(file_data)
                file_size = len(file_bytes)
                
                return {
                    "file_data": file_data,
                    "mime_type": mime_type,
                    "file_size": file_size,
                    "file_name": None,
                    "content": content
                }
            else:
                # It's plain text content
                return {
                    "content": content,
                    "mime_type": "text/plain",
                    "file_size": len(content.encode('utf-8')),
                    "file_name": None
                }
                
        except Exception as e:
            raise ValueError(f"Erro ao processar dados de documento: {e}")
    
    async def process(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Processa documento e retorna sumarização"""
        try:
            # Extract document data from content field
            if 'content' in data:
                # The document data is in the content field
                data = self._prepare_document_data(data['content'])
            elif 'file_data' in data and ('mime_type' not in data or 'file_size' not in data):
                # Fallback for direct file_data
                data = self._prepare_document_data(data['file_data'])
            
            # Initialize variables
            temp_file_path = None
            file_extension = None
            
            # Validate file size if file data is provided
            if 'file_size' in data and data['file_size']:
                self.validate_file_size(data['file_size'], self.config.max_document_size_bytes)
            
            # Validate MIME type if provided
            if 'mime_type' in data and data['mime_type']:
                self.validate_mime_type(data['mime_type'], self.allowed_mime_types)
                file_extension = self._get_file_extension(data['mime_type'], data.get('file_name'))
            
            # Decode and save document file if file data is provided
            if 'file_data' in data and data['file_data'] and file_extension:
                temp_file_path = self.decode_base64_data(data['file_data'], file_extension)
            
            try:
                # Extract text from document
                if temp_file_path and file_extension:
                    document_text = await self._extract_text(temp_file_path, file_extension)
                else:
                    # If no file data, use content directly as text
                    document_text = data.get('content', '')
                
                # Summarize using Ollama Gemma
                result = await self._summarize_with_ollama(document_text, data)
                
                # Convert to dict
                return result
                
            finally:
                # Cleanup temp file if it exists
                if temp_file_path:
                    self.cleanup_temp_file(temp_file_path)
                
        except Exception as e:
            print(f"❌ Erro no processamento de documento: {e}")
            raise
    
    def _get_file_extension(self, mime_type: str, filename: str = None) -> str:
        """Obtém extensão do arquivo"""
        # Try to get from filename first
        if filename and '.' in filename:
            return filename.split('.')[-1].lower()
        
        # Map MIME types to extensions
        mime_to_ext = {
            "application/pdf": "pdf",
            "application/msword": "doc",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
            "application/vnd.ms-excel": "xls",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
            "text/plain": "txt",
            "text/csv": "csv"
        }
        
        return mime_to_ext.get(mime_type, "txt")
    
    async def _extract_text(self, file_path: str, file_extension: str) -> str:
        """Extrai texto do documento"""
        try:
            print(f"📄 Extraindo texto do documento: {file_extension}")
            
            if file_extension == "pdf":
                return self._extract_pdf_text(file_path)
            elif file_extension in ["doc", "docx"]:
                return self._extract_word_text(file_path)
            elif file_extension in ["xls", "xlsx"]:
                return self._extract_excel_text(file_path)
            elif file_extension in ["txt", "csv"]:
                return self._extract_text_file(file_path)
            else:
                raise ValueError(f"Tipo de arquivo não suportado: {file_extension}")
                
        except Exception as e:
            print(f"❌ Erro na extração de texto: {e}")
            raise ValueError(f"Erro na extração de texto: {e}")
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extrai texto de PDF"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() + "\n"
                
                return text.strip()
        except Exception as e:
            raise ValueError(f"Erro ao extrair texto do PDF: {e}")
    
    def _extract_word_text(self, file_path: str) -> str:
        """Extrai texto de documento Word"""
        try:
            doc = Document(file_path)
            text = ""
            
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Erro ao extrair texto do Word: {e}")
    
    def _extract_excel_text(self, file_path: str) -> str:
        """Extrai texto de planilha Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Planilha: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                
                text += "\n"
            
            return text.strip()
        except Exception as e:
            raise ValueError(f"Erro ao extrair texto do Excel: {e}")
    
    def _extract_text_file(self, file_path: str) -> str:
        """Extrai texto de arquivo de texto simples"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
        except UnicodeDecodeError:
            # Try with different encoding
            with open(file_path, 'r', encoding='latin-1') as file:
                return file.read().strip()
        except Exception as e:
            raise ValueError(f"Erro ao extrair texto do arquivo: {e}")
    
    async def _summarize_with_ollama(self, document_text: str, summarize_data: Dict[str, Any]) -> Dict[str, Any]:
        """Sumariza documento usando Ollama Gemma"""
        try:
            print(f"📝 Iniciando sumarização com Ollama Gemma...")
            
            # Truncate text if too long
            max_chars = self.config.max_text_length
            if len(document_text) > max_chars:
                document_text = document_text[:max_chars] + "..."
                print(f"⚠️ Texto truncado para {max_chars} caracteres")
            
            # Ollama API request
            ollama_url = f"{self.config.ollama_base_url}/api/generate"
            
            prompt = f"""Analise o seguinte documento e crie um resumo estruturado em português brasileiro.

Documento:
{document_text}

Por favor, forneça:
1. Um resumo conciso do conteúdo principal
2. Os pontos-chave mais importantes (lista com até 8 itens)
3. O tipo de documento identificado

Formate a resposta como:
RESUMO: [resumo do conteúdo]
PONTOS-CHAVE:
- [ponto 1]
- [ponto 2]
- ...
TIPO: [tipo do documento]"""
            
            payload = {
                "model": self.config.ollama_model_resumo,
                "prompt": prompt,
                "stream": False
            }
            
            async with httpx.AsyncClient(timeout=180.0) as client:
                response = await client.post(ollama_url, json=payload)
                response.raise_for_status()
                
                result = response.json()
                summary_text = result.get("response", "").strip()
                
                if not summary_text:
                    raise ValueError("Sumarização vazia retornada pelo modelo")
                
                # Parse the structured response
                parsed_result = self._parse_summary(summary_text)
                
                print(f"✅ Sumarização concluída")
                
                return {
                    "summary": parsed_result["summary"],
                    "key_points": parsed_result["key_points"],
                    "document_type": parsed_result["document_type"],
                    "confidence": parsed_result["confidence"],
                    "success": True
                }
                
        except httpx.HTTPError as e:
            print(f"❌ Erro HTTP na sumarização: {e}")
            raise ValueError(f"Erro na comunicação com Ollama: {e}")
        except Exception as e:
            print(f"❌ Erro na sumarização: {e}")
            raise ValueError(f"Erro na sumarização: {e}")
    
    def _parse_summary(self, summary_text: str) -> Dict[str, Any]:
        """Parse a sumarização estruturada"""
        try:
            lines = summary_text.split('\n')
            
            summary = ""
            key_points = []
            document_type = "documento"
            
            current_section = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                if line.startswith("RESUMO:"):
                    summary = line.replace("RESUMO:", "").strip()
                    current_section = "summary"
                elif line.startswith("PONTOS-CHAVE:"):
                    current_section = "keypoints"
                elif line.startswith("TIPO:"):
                    document_type = line.replace("TIPO:", "").strip()
                    current_section = None
                elif current_section == "keypoints" and line.startswith("-"):
                    key_point = line.replace("-", "").strip()
                    if key_point:
                        key_points.append(key_point)
                elif current_section == "summary" and not summary:
                    summary = line
            
            # Fallback if parsing failed
            if not summary:
                summary = summary_text[:500] + "..." if len(summary_text) > 500 else summary_text
            
            if not key_points:
                # Try to extract key points from the text
                sentences = summary.split('.')
                key_points = [s.strip() for s in sentences[:5] if s.strip()]
            
            # Estimate confidence
            confidence = min(0.95, max(0.7, len(summary) / 200))
            
            return {
                "summary": summary,
                "key_points": key_points[:8],  # Limit to 8 points
                "document_type": document_type,
                "confidence": confidence
            }
            
        except Exception as e:
            print(f"⚠️ Erro ao fazer parse da sumarização: {e}")
            # Fallback
            return {
                "summary": summary_text[:500] + "..." if len(summary_text) > 500 else summary_text,
                "key_points": [],
                "document_type": "documento",
                "confidence": 0.7
            }
